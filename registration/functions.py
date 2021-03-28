from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from django.views.static import serve
import openpyxl as xl
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from .models import DriverInfo, City
from twilio.rest import Client
import random, string


def send_sms(token):
    account_sid = 'ACe67ca4c930cd998ecc8f766d6af34ed8'
    auth_token = '50877af1f1f73b5dd20d74b8a8e2cab7'
    client = Client(account_sid, auth_token)

    message = client.messages.create(
        body=f" کد ثبت نام {token} ",
        from_='+14159445765',
        to='+989126753305'
    )
    print(message.sid)


def create_auth_code():
    return "".join(random.choices(string.digits, k=6))


def render_pdf_view(request):
    drivers = DriverInfo.objects.all()
    template_path = 'list_pdf.html'
    context = {'drivers': drivers}
    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = "inline;filename = list.pdf"
    # response['Content-Transfer-Encoding'] = 'binary'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(html, dest=response, )
    # if error then show some funy view
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def export_excel(request):
    # creating excel file
    wb = xl.Workbook()
    ws1 = wb.create_sheet('drivers', 0)
    drivers = DriverInfo.objects.all()
    ws1.append(
        ['کد ملی', 'کارت هوشمند', 'نام', 'نام خانوادگی', 'شماره همراه', 'آدرس ایمیل',
         'تاریخ ثبت نام', 'شهر', 'استان'])
    for rows in ws1.iter_rows(min_row=1, max_row=1, min_col=1, max_col=9):
        for cell in rows:
            cell.fill = PatternFill(start_color='c9e1f8', end_color='c9e1f8', fill_type='solid')
    ws1.sheet_properties.tabColor = "1072BA"
    for driver in drivers:
        ws1.append([driver.national_id, driver.smart_card_number, driver.name, driver.family,
                    driver.cellnumber, driver.email, driver.date_added, driver.city.name, driver.state.name])
    # creating response object
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;filename = "Drivers.xlsx"'
    wb.save(response)
    return response
    # wb.save('Drivers.xlsx')
    # serve method khode django
    # filepath = "C:/Users/farhood/PycharmProjects/pythonProject/RegService/drivers.xlsx"
    # return serve(request, os.path.basename(filepath), os.path.dirname(filepath))


